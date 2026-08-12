#!/usr/bin/env python3
"""
The Running Postman — wine list migration   [RETIRED — DO NOT RUN]
==================================================================

*** This has already been run. The live Google Sheet is now the source of  ***
*** truth. Running it again would rebuild a sheet from the original messy  ***
*** workbook and throw away everything the venue has edited since.         ***

Kept because the next hospitality client will arrive with the same kind of
spreadsheet — section banners typed into cells, wines duplicated to act as
flags, prices in the vintage column — and this already solves that. Treat it as
a starting point to adapt, not something to execute against a live sheet.

Reads the exported workbook and produces a single clean, sortable spreadsheet.

    python3 scripts/migrate_sheet.py <input.xlsx> <output.xlsx>

What it does
------------
* keeps only the consolidated `APP Link` data and drops the 15 presentation tabs
* splits Producer out of Name, and the grape list out of the trailing " - "
* folds blend continuation rows ("- Cabernet/Merlot") into the wine above
* removes blank spacer rows and section banners typed into random columns
* rejects the fortified section's price-in-the-Vintage-column values
* splits "Clare Valley, SA" into Region + Country
* collapses the duplicated Organic / Vegan / Cellar listings into tick boxes
* pulls glass prices across from the `BY THE GLASS.` tab
* assigns a stable ID to every wine
* flags anything ambiguous on a Migration Review tab instead of guessing

The source workbook is only ever read, never modified.
"""

import json
import os
import re
import sys
import unicodedata
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

SOURCE_TAB = "APP Link"
GLASS_TAB = "BY THE GLASS."
OUTPUT_TAB = "Wine List"
REVIEW_TAB = "Migration Review"
LISTS_TAB = "_Lists"

HEADERS = [
    "ID", "Producer", "Wine", "Vintage", "Section", "Variety", "Country", "Region",
    "Price Bottle", "Price Glass", "Available", "Organic", "Vegan", "Biodynamic",
    "Cellar List", "Featured",
    # Tasting data. These drive Find My Wine — a wine without them still appears
    # on the list, but the app has to fall back to a guess based on its variety.
    "Body", "Sweetness", "Acidity", "Tannins", "Pairing Tags",
    "short_desc", "long_desc",
    "Notes",
]

# Deliberately no stock columns. The venue already has a stocktake system, and a
# half-used Qty / Par / Last Counted set here would just become a second, wrong
# answer to "how many have we got". `Available` is the only stock-ish field: it
# controls whether a wine shows in the app, nothing more.

# Where hand-written tasting data was merged from during the original run. That
# file has since been removed — the live sheet owns this data now — so the merge
# step simply no-ops. Left in place because the pattern is worth reusing for the
# next venue.
NOTES_FILE = "data/tasting-notes.json"

SECTION_MAP = {
    "BUBBLES": "Bubbles",
    "WHITE": "White",
    "ITALIAN WHITE": "Italian White",
    "ROSE": "Rosé",
    "ROSÉ": "Rosé",
    "CHILLED RED": "Chilled Red",
    "RED": "Red",
    "ITALIAN REDS": "Italian Red",
    "ITALIAN RED": "Italian Red",
    "FORTIFIED": "Fortified & Sweet",
}

SECTIONS = [
    "Bubbles", "White", "Italian White", "Rosé",
    "Chilled Red", "Red", "Italian Red", "Fortified & Sweet",
]

# Menu order, not alphabetical — an A-Z sort would put Chilled Red second and
# Bubbles nowhere near the front.
SECTION_RANK = {name: n for n, name in enumerate(SECTIONS)}

COUNTRIES = [
    "Australia", "New Zealand", "France", "Italy", "Spain", "Germany",
    "Austria", "Portugal", "Argentina", "USA", "South Africa", "Chile",
]

# Region suffix -> country. Recovers the ~100 rows with a region but no country.
SUFFIX_COUNTRY = {
    "VIC": "Australia", "NSW": "Australia", "SA": "Australia", "WA": "Australia",
    "TAS": "Australia", "QLD": "Australia", "NT": "Australia", "ACT": "Australia",
    "AUS": "Australia", "AUSTRALIA": "Australia",
    "NZ": "New Zealand",
    "FRA": "France", "FRANCE": "France",
    "ITA": "Italy", "ITALY": "Italy",
    "GER": "Germany",
    "ESP": "Spain", "SPN": "Spain", "EPN": "Spain",  # EPN is a recurring typo
    "ARG": "Argentina", "ARGENTINA": "Argentina",
    "USA": "USA",
    "AU": "Austria", "AUT": "Austria", "AUSTRIA": "Austria",
    "PRT": "Portugal", "PORTUGAL": "Portugal",
    "SPAIN": "Spain", "GERMANY": "Germany", "NZ.": "New Zealand",
}

BANNERS = {
    "white", "red", "rose", "rosé", "bubbles", "italian red", "italian white",
    "organic", "vegan", "cellar", "dessert/ fortified", "dessert / fortified",
    "barolo", "nebbiolo", "valpolicella", "montepulciano", "whites/rose",
    "reds", "shiraz and cabernet",
}

# Words too generic to identify a wine by. Used to stop the by-the-glass matcher
# pairing "Chianti Classico" with an unrelated Chianti in the main list.
GENERIC = {
    "sparkling", "rose", "rosé", "white", "red", "house", "premium", "blend",
    "sauvignon", "blanc", "chardonnay", "riesling", "pinot", "gris", "noir",
    "shiraz", "chianti", "classico", "botrytis", "viognier", "grenache",
    "montepulciano", "fiano", "orange", "chilled", "field", "etna", "bianco",
    "chablis", "malbec", "nero", "avola", "italian", "australian", "carafe",
}


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def txt(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def money(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    cleaned = re.sub(r"[^0-9.]", "", str(v))
    if not cleaned:
        return None
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return None


def is_true(v):
    return str(v).strip().lower() in {"true", "yes", "y", "1", "x"}


def clean_vintage(v):
    """A year, NV or MV. Everything else — like the prices sitting in the
    fortified section's Vintage column — is rejected."""
    s = txt(v).upper()
    if s.endswith(".0"):           # Excel turns 2014 into 2014.0
        s = s[:-2]
    if s in {"NV", "MV"}:
        return s
    if re.fullmatch(r"\d{4}", s) and 1900 <= int(s) <= 2035:
        return s
    return ""


def split_name(raw):
    """'Louis Roederer, 'Collection 244' - Pinot Noir/Chardonnay'
       -> producer, wine, grapes"""
    if not raw:
        return "", "", ""
    working, grapes = raw, ""
    m = re.search(r"\s[-–—]\s", working)
    if m:
        grapes = txt(working[m.end():])
        working = working[:m.start()]
    if "," in working:
        head, _, tail = working.partition(",")
        return txt(head), txt(tail), grapes
    return "", txt(working), grapes


def split_region(raw):
    """'Clare Valley, SA' -> ('Clare Valley', 'Australia')"""
    s = txt(raw)
    if not s:
        return "", ""
    if "," not in s:
        return s, ""
    parts = s.split(",")
    suffix = re.sub(r"[.\s]", "", parts[-1]).upper()
    country = SUFFIX_COUNTRY.get(suffix, "")
    if not country:
        return s, ""
    return txt(",".join(parts[:-1])), country


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


def tokens(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return {t for t in re.split(r"[^a-z0-9]+", s.lower()) if len(t) > 3}


# --------------------------------------------------------------------------- #
# Read
# --------------------------------------------------------------------------- #

def find_header_row(ws, must_contain="Category"):
    """The export has a blank first row, so the header isn't necessarily row 1.
    Find it rather than assuming — this bites the live CSV fetch too."""
    for r in range(1, min(ws.max_row, 20) + 1):
        values = [txt(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if must_contain in values:
            return r
    raise SystemExit(f"Could not find a header row containing {must_contain!r}")


def read_wines(ws, review):
    header_row = find_header_row(ws)
    headers = {txt(ws.cell(header_row, c).value): c
               for c in range(1, ws.max_column + 1)
               if txt(ws.cell(header_row, c).value)}

    def cell(r, name):
        c = headers.get(name)
        return ws.cell(r, c).value if c else None

    wines = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw_name = txt(cell(r, "Name"))
        price_bottle = money(cell(r, "Price Bottle"))
        price_glass = money(cell(r, "Price Glass"))
        has_price = price_bottle is not None or price_glass is not None
        section = SECTION_MAP.get(txt(cell(r, "Category")).upper(), "")

        # Blend continuation line — belongs to the wine above it.
        if raw_name and not has_price and re.match(r"^[-–—]", raw_name):
            if wines and not wines[-1]["variety"]:
                wines[-1]["variety"] = txt(re.sub(r"^[-–—]\s*", "", raw_name))
            continue

        if not raw_name:
            continue

        # Section banner typed into the Name or Vintage column.
        vintage_raw = txt(cell(r, "Vintage"))
        if not has_price and not section and raw_name.lower() in BANNERS:
            continue
        if not has_price and not raw_name and vintage_raw.lower() in BANNERS:
            continue

        if not section:
            if has_price:
                review.append([r, raw_name, "Blank or unrecognised Category",
                               "Row has a price but no section, so it would never "
                               "appear in the app. Set Category and re-run."])
            continue

        if not has_price:
            review.append([r, raw_name, "No price",
                           "Skipped — add a bottle or glass price."])
            continue

        producer, wine_name, grapes = split_name(raw_name)
        raw_region = txt(cell(r, "Region"))
        region, implied_country = split_region(raw_region)
        stated_country = txt(cell(r, "Country"))

        # Where Country and the region suffix disagree, flag rather than guess —
        # these are genuine errors and only a human knows which side is right.
        country = stated_country
        if stated_country and implied_country and stated_country != implied_country:
            review.append([r, raw_name, "Country contradicts Region",
                           f'Country says "{stated_country}" but Region "{raw_region}" '
                           f'implies {implied_country}. Kept "{stated_country}" — please confirm.'])
        elif not stated_country:
            country = implied_country

        if not country:
            review.append([r, raw_name, "No country",
                           "Could not determine a country. Please fill it in."])

        vintage = clean_vintage(cell(r, "Vintage"))
        if vintage_raw and not vintage:
            review.append([r, raw_name, "Vintage is not a year",
                           f'Vintage cell contained "{vintage_raw}" — cleared. '
                           f"Set a year, NV or MV."])

        wines.append({
            "producer": producer,
            "wine": wine_name or raw_name,
            "vintage": vintage,
            "section": section,
            "variety": txt(cell(r, "Variety")) or grapes,
            "country": country,
            "region": region,
            "price_bottle": price_bottle,
            "price_glass": price_glass,
            "organic": is_true(cell(r, "Organic")),
            "vegan": is_true(cell(r, "Vegan")),
            "cellar": is_true(cell(r, "Cellar")),
            "featured": is_true(cell(r, "Featured")),
            "source_row": r,
        })

    return wines


# The glass menu's own section banners, so wines pulled across land in the right
# part of the list instead of arriving sectionless.
GLASS_BANNERS = {
    "SPARKLING": "Bubbles",
    "WHITE": "White",
    "REDS": "Red",
    "RED": "Red",
    "DESSERT": "Fortified & Sweet",
    "FORTIFIED": "Fortified & Sweet",
}


# Menu-speak that isn't a grape. "Premium Riesling" is a Riesling; the word only
# tells you where it sits on the printed list.
VARIETY_PREFIXES = ("premium ", "australian ", "italian ", "house ")


def read_glass(ws):
    """The by-the-glass menu is a separate curated list, laid out for printing.

    The layout is strictly positional, and reading it that way matters: an
    earlier version picked "the first string longer than 12 characters" as the
    wine name, which silently grabbed the *variety* for anything called
    "Sauvignon Blanc" or "Premium Chardonnay" and threw the producer away.

        c1 vintage   c2 variety   c3 producer/cuvée - region   c4 glass   c5 bottle
    """
    out = []
    section = ""
    carafe = False       # the House pours are priced per 500ml carafe, not per bottle

    for r in range(1, ws.max_row + 1):
        c1, c2, c3, c4, c5 = (ws.cell(r, c).value for c in range(1, 6))

        # Measure header row: "150ml" / "BTL" or "Carafe 500ml".
        if isinstance(c4, str) and "ml" in c4.lower():
            carafe = isinstance(c5, str) and "carafe" in c5.lower()
            continue

        # Section banner: text in c1 with nothing else on the row.
        if isinstance(c1, str) and c1.strip() and not any([c2, c3, c4, c5]):
            head = c1.upper().rstrip(".")
            if "ROSE" in head or "ORANGE" in head:
                section = "Rosé"       # refined per-wine below
            else:
                section = next((v for k, v in GLASS_BANNERS.items() if k in head), section)
            continue

        if not isinstance(c4, (int, float)):
            continue

        variety = txt(c2)
        name = txt(c3)
        if not (variety or name):
            continue

        # House White / House Red have no producer — the variety cell is the name.
        display = name or variety

        low = variety.lower()
        for p in VARIETY_PREFIXES:
            if low.startswith(p):
                variety = variety[len(p):].strip()
                break

        wine_section = section
        if section == "Rosé":
            if "orange" in low:
                wine_section = "White"
            elif "chilled" in low or low == "red":
                wine_section = "Chilled Red"

        out.append({
            "raw": display,
            "variety": variety,
            "vintage": clean_vintage(c1),
            "glass": round(float(c4), 2),
            "bottle": round(float(c5), 2) if isinstance(c5, (int, float)) and not carafe else None,
            "carafe": round(float(c5), 2) if isinstance(c5, (int, float)) and carafe else None,
            "section": wine_section,
            "row": r,
        })
    return out


# --------------------------------------------------------------------------- #
# Dedupe
# --------------------------------------------------------------------------- #

def dedupe(wines, review):
    """The sheet flags organic / vegan / cellar wines by repeating the whole wine
    in a trailing section, so the same bottle can appear three times."""
    merged = OrderedDict()

    for w in wines:
        key = norm(f"{w['producer']}|{w['wine']}|{w['vintage']}")
        if key not in merged:
            merged[key] = w
            continue

        first = merged[key]
        first["organic"] = first["organic"] or w["organic"]
        first["vegan"] = first["vegan"] or w["vegan"]
        first["featured"] = first["featured"] or w["featured"]
        first["variety"] = first["variety"] or w["variety"]
        first["country"] = first["country"] or w["country"]
        if first["price_glass"] is None:
            first["price_glass"] = w["price_glass"]

        # `cellar` is deliberately NOT merged. It says which list a bottle sits
        # on, and several wines are on both the current list and the cellar
        # list. Merging would mark those cellar-only and hide them from what's
        # actually pouring. First row wins; the main list comes first.

        if (w["price_bottle"] and first["price_bottle"]
                and w["price_bottle"] != first["price_bottle"]):
            review.append([
                w["source_row"], f"{w['producer']} {w['wine']} {w['vintage']}".strip(),
                "Duplicate with a different price",
                f"Also on row {first['source_row']} at ${first['price_bottle']:.0f} "
                f"(this one ${w['price_bottle']:.0f}). Kept ${first['price_bottle']:.0f}.",
            ])

    return list(merged.values())


def apply_tasting_notes(wines, path, review):
    """Merge the hand-written tasting data in, matching on producer + wine.

    Matching is loose on purpose — punctuation and curly quotes in the sheet
    change often enough that an exact match would silently orphan entries. Any
    entry that fails to find a home is reported rather than quietly dropped."""
    if not os.path.exists(path):
        return 0

    with open(path, encoding="utf-8") as fh:
        entries = json.load(fh).get("wines", [])

    index = {}
    by_wine = {}
    for w in wines:
        index.setdefault(norm(f"{w['producer']}{w['wine']}"), []).append(w)
        by_wine.setdefault(norm(w["wine"]), []).append(w)

    applied = 0
    for e in entries:
        key = norm(f"{e.get('producer', '')}{e.get('wine', '')}")
        candidates = index.get(key, [])

        # Some rows never had a comma in the original name, so the producer
        # couldn't be split out and sits empty. Fall back to matching on the
        # wine name alone.
        if not candidates:
            candidates = by_wine.get(norm(e.get("wine", "")), [])

        # Several vintages of the same wine can share a key.
        if e.get("vintage"):
            candidates = [c for c in candidates if c["vintage"] == e["vintage"]] or candidates

        if not candidates:
            review.append(["—", f"{e.get('producer', '')} {e.get('wine', '')}".strip(),
                           "Tasting notes didn't match any wine",
                           f"An entry in {NOTES_FILE} found no matching wine on the list. "
                           f"Check the producer and wine name still line up."])
            continue

        target = candidates[0]

        # If the sheet couldn't work out the producer, the notes file knows it.
        # Fill it in and trim the duplicated prefix off the wine name.
        if not target["producer"] and e.get("producer"):
            target["producer"] = e["producer"]
            if target["wine"].lower().startswith(e["producer"].lower()):
                target["wine"] = target["wine"][len(e["producer"]):].strip(" ,-")

        target["body"] = e.get("body")
        target["sweetness"] = e.get("sweetness")
        target["acidity"] = e.get("acidity")
        target["tannins"] = e.get("tannins")
        target["pairing"] = ", ".join(e.get("pairing", []))
        target["short_desc"] = e.get("short_desc", "")
        target["long_desc"] = e.get("long_desc", "")

        # The Bubbles block had "Coonawarra, SA" filled down over Tasmanian and
        # Italian producers, so the notes file can carry corrections.
        if e.get("region"):
            target["region"] = e["region"]
        if e.get("country"):
            target["country"] = e["country"]

        applied += 1

    return applied


def apply_glass_prices(wines, glass, review):
    """Match the by-the-glass menu onto the main list where we can be confident,
    and add the rest as new rows.

    Matching has to be strict. One shared word is not enough: "Georges Vesselle"
    (Champagne, $130) and "Georges, 'The Tiller'" (Clare Valley Shiraz, $85)
    share the word "Georges", and a single-token match silently puts a $21 glass
    price on a Champagne.

    So a match needs two distinctive words in common — and the bottle prices must
    be in the same ballpark. The price acts as a veto: Soul Growers appears twice
    with two shared words ("Soul", "Growers") but $60 against $195, which is
    plainly a different wine by the same producer.

    Anything not confidently matched is added as its own row and flagged, because
    a wrong price on the menu is worse than a duplicate someone can merge."""
    index = [(tokens(f"{w['producer']} {w['wine']}") - GENERIC, w) for w in wines]
    added = 0

    # The glass menu writes regions without a state suffix ("Mclaren Vale" rather
    # than "Mclaren Vale, SA"), so there's nothing to derive a country from.
    # The main list already knows where these regions are — reuse that instead of
    # hardcoding a gazetteer.
    region_country = {}
    for w in wines:
        if w["region"] and w["country"]:
            region_country.setdefault(norm(w["region"]), w["country"])

    for g in glass:
        head = g["raw"].split(" - ")[0]
        keys = tokens(head) - GENERIC

        # Producers list the same cuvée in several varieties — Holm Oak
        # 'Protégé' is both a Chardonnay and a Pinot Noir — so the variety on
        # the glass menu is used as a guard, not just as a label.
        glass_variety = tokens(g["variety"])

        # Score every candidate, then take the best — not the first one seen.
        best, best_score = None, 0
        if keys:
            for toks, w in index:
                shared = len(keys & toks)
                if shared < 2:
                    continue

                # Price veto: same producer, wildly different price = different wine.
                if w["price_bottle"] and g["bottle"]:
                    ratio = abs(w["price_bottle"] - g["bottle"]) / max(w["price_bottle"], g["bottle"])
                    if ratio > 0.15:
                        continue

                # Variety veto: if both sides name a variety and they share
                # nothing, it's a different wine from the same producer.
                cand_variety = tokens(w["variety"])
                if glass_variety and cand_variety and not (glass_variety & cand_variety):
                    continue

                if shared > best_score:
                    best, best_score = w, shared

        match = best

        if match:
            match["price_glass"] = g["glass"]
            if match["price_bottle"] and g["bottle"] and abs(match["price_bottle"] - g["bottle"]) > 2:
                review.append([
                    g["row"], g["raw"], "Glass menu bottle price differs",
                    f"Matched to {match['producer']} {match['wine']}. By the glass tab "
                    f"says ${g['bottle']:.0f}, main list says ${match['price_bottle']:.0f}. "
                    f"Kept the main list price — confirm the match is right.",
                ])
        else:
            # On this tab the text after " - " is the region, not the grapes —
            # the opposite convention to the main list.
            head, _, tail = g["raw"].partition(" - ")
            region_clean = txt(tail)
            # Strip bottle-size notes like "(500ml)" that ride along in the region.
            region_clean = re.sub(r"\s*\(\s*\d+\s*ml\s*\)?\s*$", "", region_clean, flags=re.I)
            region_clean, country = split_region(region_clean)
            if not country:
                country = region_country.get(norm(region_clean), "")

            producer, wine_name, _ = split_name(head)
            if not producer:
                # No comma, so the whole thing is the producer ("Settlement Road")
                # and the variety column carries the wine's identity.
                producer = txt(head)
                wine_name = g["variety"] or txt(head)

            wines.append({
                "producer": producer,
                "wine": wine_name,
                "vintage": g["vintage"],
                "section": g.get("section", ""),
                "variety": g["variety"],
                "country": country,
                "region": region_clean,
                "price_bottle": g["bottle"],
                "price_glass": g["glass"],
                "organic": False, "vegan": False, "cellar": False, "featured": False,
                "source_row": f"glass!{g['row']}",
            })
            added += 1

            # Only surface these when something is actually wrong with them.
            # A glass wine that came across with a producer, a region and a
            # country needs no human attention, and 25 "FYI" rows just bury the
            # six items that genuinely do.
            missing = []
            if not producer:
                missing.append("no producer could be read from the name")
            if not region_clean:
                missing.append("no region")
            if not country:
                missing.append("no country")

            if missing:
                note = (f"Added from the By The Glass tab into {g.get('section') or 'no section'}, "
                        f"but {', and '.join(missing)}. Please fill in.")
                if g.get("carafe"):
                    note += (f" Priced by the 500ml carafe (${g['carafe']:.0f}) rather than "
                             f"by the bottle, so Price Bottle is blank.")
                review.append([
                    f"glass!{g['row']}", f"{producer} — {wine_name}".strip(" —"),
                    "Added from By The Glass tab — incomplete", note,
                ])

    return added


# --------------------------------------------------------------------------- #
# Write
# --------------------------------------------------------------------------- #

HEADER_FILL = PatternFill("solid", fgColor="221C18")
HEADER_FONT = Font(bold=True, color="F6EFE6")
AMBER = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="F8D7DA")
GREEN = PatternFill("solid", fgColor="E3F2E1")


def write_output(path, wines, review):
    wb = openpyxl.Workbook()

    # ---- Wine List ----
    ws = wb.active
    ws.title = OUTPUT_TAB
    ws.append(HEADERS)

    for i, w in enumerate(wines, start=1):
        ws.append([
            f"RP-{i:04d}", w["producer"], w["wine"], w["vintage"], w["section"],
            w["variety"], w["country"], w["region"],
            w["price_bottle"], w["price_glass"],
            True,                                   # Available
            w["organic"], w["vegan"], False,        # Biodynamic — nothing to migrate
            w["cellar"], w["featured"],
            w.get("body"), w.get("sweetness"), w.get("acidity"), w.get("tannins"),
            w.get("pairing", ""), w.get("short_desc", ""), w.get("long_desc", ""),
            None,                                   # Notes
        ])

    last = ws.max_row
    for c, head in enumerate(HEADERS, start=1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    # Vintage as text, or Excel turns 2014 into a number and NV into junk.
    for r in range(2, last + 1):
        ws.cell(r, 4).number_format = "@"
        ws.cell(r, 9).number_format = '"$"#,##0.00'
        ws.cell(r, 10).number_format = '"$"#,##0.00'

    widths = {1: 9, 2: 26, 3: 42, 4: 9, 5: 17, 6: 22, 7: 14, 8: 22,
              9: 13, 10: 12, 11: 11, 12: 10, 13: 9, 14: 12, 15: 12, 16: 11,
              17: 8, 18: 11, 19: 9, 20: 9, 21: 30, 22: 60, 23: 90, 24: 30}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(2, last + 1):
        ws.cell(r, 23).alignment = Alignment(wrap_text=True, vertical="top")
    # Deliberately no auto-filter. Google Sheets imports it as a filter, and a
    # filter overlapping the data blocks Format -> Convert to table ("please
    # remove the filter that overlaps with the conversion area"). The table
    # provides better per-column filtering anyway.

    # ---- Lists ----
    lists = wb.create_sheet(LISTS_TAB)
    lists["A1"], lists["B1"], lists["C1"] = "Section", "Country", "Variety"
    for c in "ABC":
        lists[f"{c}1"].fill, lists[f"{c}1"].font = HEADER_FILL, HEADER_FONT

    varieties = sorted({w["variety"] for w in wines if w["variety"]})
    for i, v in enumerate(SECTIONS, start=2):
        lists.cell(i, 1, v)
    for i, v in enumerate(COUNTRIES, start=2):
        lists.cell(i, 2, v)
    for i, v in enumerate(varieties, start=2):
        lists.cell(i, 3, v)
    for c, w in {1: 20, 2: 18, 3: 30}.items():
        lists.column_dimensions[get_column_letter(c)].width = w

    # Dropdowns. Section and Country reject bad input; Variety only warns, so a
    # genuinely new grape can still be entered and tidied later.
    def add_dv(col, ref, allow_blank, strict):
        dv = DataValidation(type="list", formula1=ref, allow_blank=allow_blank,
                            showErrorMessage=strict)
        dv.error = "Pick a value from the list, or add it to the _Lists tab first."
        dv.errorTitle = "Not on the list"
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max(last, 2)}")

    add_dv("E", f"'{LISTS_TAB}'!$A$2:$A${len(SECTIONS) + 1}", False, True)
    add_dv("G", f"'{LISTS_TAB}'!$B$2:$B${len(COUNTRIES) + 1}", False, True)
    add_dv("F", f"'{LISTS_TAB}'!$C$2:$C${len(varieties) + 1}", True, False)

    # ---- Conditional formatting: make invisible problems visible ----
    body = f"A2:{get_column_letter(len(HEADERS))}{max(last, 2)}"
    ws.conditional_formatting.add(body, FormulaRule(
        formula=['AND($C2<>"",$I2="",$J2="")'], fill=AMBER, stopIfTrue=False))
    ws.conditional_formatting.add(body, FormulaRule(
        formula=['AND($C2<>"",OR($E2="",$F2="",$G2=""))'], fill=AMBER, stopIfTrue=False))
    ws.conditional_formatting.add(body, FormulaRule(
        formula=['AND($C2<>"",COUNTIFS($B:$B,$B2,$C:$C,$C2,$D:$D,$D2)>1)'],
        fill=RED, stopIfTrue=False))
    # Green = has real tasting data, so Find My Wine isn't guessing for it.
    ws.conditional_formatting.add(body, FormulaRule(
        formula=['AND($C2<>"",$Q2<>"")'], fill=GREEN, stopIfTrue=False))

    # ---- Migration Review ----
    rv = wb.create_sheet(REVIEW_TAB)
    rv.append(["Source row", "Wine", "Issue", "What to do"])
    for c in range(1, 5):
        rv.cell(1, c).fill = PatternFill("solid", fgColor="8C2F39")
        rv.cell(1, c).font = Font(bold=True, color="FFFFFF")
    for row in review:
        rv.append(row)
    rv.freeze_panes = "A2"
    for c, w in {1: 12, 2: 44, 3: 32, 4: 78}.items():
        rv.column_dimensions[get_column_letter(c)].width = w
    for r in range(2, rv.max_row + 1):
        rv.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


# --------------------------------------------------------------------------- #

def main():
    if len(sys.argv) != 3:
        raise SystemExit(__doc__)
    src, dest = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(src, data_only=True)
    review = []

    wines = read_wines(wb[SOURCE_TAB], review)
    raw_count = len(wines)

    wines = dedupe(wines, review)
    deduped = raw_count - len(wines)

    noted = apply_tasting_notes(wines, NOTES_FILE, review)

    added = 0
    if GLASS_TAB in wb.sheetnames:
        added = apply_glass_prices(wines, read_glass(wb[GLASS_TAB]), review)

    # Sort into menu order before writing, so the sheet reads like the printed
    # list and IDs run in the same order.
    #
    # The source was fragmented: the cellar back-vintages sat in their own block
    # at the end spanning several sections, and the by-the-glass wines were
    # appended after everything. That left Whites near the top AND the bottom.
    #
    # Within a section, current wines come before cellar ones, then group by
    # variety (all the Rieslings together, as the printed menu does), then
    # ascending price — the order someone actually scans a wine list in.
    wines.sort(key=lambda w: (
        SECTION_RANK.get(w["section"], 99),
        1 if w["cellar"] else 0,
        (w["variety"] or "zzz").lower(),
        w["price_bottle"] if w["price_bottle"] is not None else 1e9,
        (w["producer"] or "").lower(),
    ))

    write_output(dest, wines, review)

    with_glass = sum(1 for w in wines if w["price_glass"])
    print(f"  source tabs           {len(wb.sheetnames)} → 3 (Wine List, _Lists, Migration Review)")
    print(f"  rows parsed           {raw_count}")
    print(f"  duplicates collapsed  {deduped}")
    print(f"  added from glass tab  {added}")
    print(f"  wines out             {len(wines)}")
    print(f"  with a glass price    {with_glass}")
    print(f"  on the cellar list    {sum(1 for w in wines if w['cellar'])}")
    print(f"  organic / vegan       {sum(1 for w in wines if w['organic'])} / "
          f"{sum(1 for w in wines if w['vegan'])}")
    print(f"  with tasting notes    {noted}")
    print(f"  review items          {len(review)}")


if __name__ == "__main__":
    main()
